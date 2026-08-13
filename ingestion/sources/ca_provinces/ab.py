"""Alberta baby names, from Service Alberta's open data (vital statistics).

Canada has no single federal baby-name dataset -- each province publishes its own.
Alberta is the only one included so far (its files are clean and rank is provided
directly); BC's open-data CSVs were attempted but the server drops the connection
for every non-interactive-browser client tried (curl, a headless fetch, and a
sandboxed browser navigation all failed) -- see ingestion/vendored/PROVENANCE.md.
Add more provinces the same way as this one if/when their data is reachable.

Reads from vendored snapshots in ingestion/vendored/ca/.
"""

import logging
from pathlib import Path

import openpyxl

from ingestion.common.schema import NormalizedNameRecord

logger = logging.getLogger("ingestion.ca_ab")

SOURCE = "ab_vital_stats"
REGION = "AB"
VENDORED_DIR = Path(__file__).resolve().parent.parent.parent / "vendored" / "ca"
FILES = ["ab_1980_2020.xlsx", "ab_2021_2024.xlsx"]

GENDER_MAP = {"Boy": "M", "Girl": "F"}


def _parse_file(path: Path) -> list[NormalizedNameRecord]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    for row in ws.iter_rows(min_row=3, max_col=5, values_only=True):
        rank, name, count, gender, year = row
        if not isinstance(rank, int) or not isinstance(name, str) or not name.strip():
            continue
        sex = GENDER_MAP.get(str(gender).strip())
        if sex is None or not isinstance(year, int):
            continue
        records.append(
            NormalizedNameRecord(
                name=name.strip().title(),
                sex=sex,
                year=year,
                country_code="CA",
                source=SOURCE,
                count=int(count) if isinstance(count, (int, float)) else None,
                rank=rank,
                region_code=REGION,
            )
        )
    return records


def fetch_records() -> list[NormalizedNameRecord]:
    all_records: list[NormalizedNameRecord] = []
    for filename in FILES:
        path = VENDORED_DIR / filename
        try:
            records = _parse_file(path)
            logger.info("parsed %d Alberta records from %s", len(records), filename)
            all_records.extend(records)
        except Exception:
            logger.exception("failed to parse %s", path)
    return all_records
