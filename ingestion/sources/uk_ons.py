"""UK (England & Wales) baby names, from ONS annual releases.

Reads from vendored snapshots (ingestion/vendored/gb/{year}_{sex}.xlsx) rather than
downloading live, per the project's offline-by-default ingestion design -- see
ingestion/vendored/PROVENANCE.md. ONS also renames these files inconsistently every
year (e.g. "2019boysnames.xlsx" vs "boysnames2022.xlsx" vs "2025boysbabynames.xlsx"),
so there's no stable URL pattern to auto-discover new years from anyway. To add a
year: download that year's boys/girls XLSX from ons.gov.uk and drop it in
ingestion/vendored/gb/ as "{year}_M.xlsx" / "{year}_F.xlsx".
"""

import logging
from pathlib import Path

import openpyxl

from ingestion.common.schema import NATIONAL, NormalizedNameRecord

logger = logging.getLogger("ingestion.uk_ons")

SOURCE = "ons"
VENDORED_DIR = Path(__file__).resolve().parent.parent / "vendored" / "gb"


def _find_full_list_sheet(wb):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=1, values_only=True):
            cell = row[0]
            if isinstance(cell, str) and "top names for baby" in cell.lower():
                return ws
    return None


def _find_header_row(ws, max_row=15) -> int | None:
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, max_col=2, values_only=True), start=1
    ):
        c0 = str(row[0]).strip().lower() if row[0] else ""
        c1 = str(row[1]).strip().lower() if row[1] else ""
        if c0 == "rank" and c1 == "name":
            return row_idx
    return None


def _parse_sheet(ws, year: int, sex: str) -> list[NormalizedNameRecord]:
    header_row = _find_header_row(ws)
    if header_row is None:
        logger.warning("could not find header row in sheet for year=%s sex=%s", year, sex)
        return []

    records = []
    for row in ws.iter_rows(min_row=header_row + 1, max_col=3, values_only=True):
        rank, name, count = row[0], row[1], row[2]
        if not isinstance(rank, int) or not isinstance(name, str) or not name.strip():
            continue
        records.append(
            NormalizedNameRecord(
                name=name.strip().title(),
                sex=sex,
                year=year,
                country_code="GB",
                source=SOURCE,
                count=int(count) if isinstance(count, (int, float)) else None,
                rank=rank,
                region_code=NATIONAL,
            )
        )
    return records


def fetch_records() -> list[NormalizedNameRecord]:
    all_records: list[NormalizedNameRecord] = []
    for path in sorted(VENDORED_DIR.glob("*.xlsx")):
        year_str, sex = path.stem.split("_")
        year = int(year_str)
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = _find_full_list_sheet(wb)
            if ws is None:
                logger.warning("no full-list sheet found in %s", path)
                continue
            records = _parse_sheet(ws, year, sex)
            logger.info("parsed %d ONS records for year=%s sex=%s", len(records), year, sex)
            all_records.extend(records)
        except Exception:
            logger.exception("failed to parse %s", path)
    return all_records
